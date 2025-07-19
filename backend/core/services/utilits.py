import os
import re
import threading
import uuid
from datetime import datetime
from io import BytesIO

from flask import current_app
from flask_mail import Message
from itsdangerous import URLSafeTimedSerializer
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename

from backend.core import mail
from backend.core.config import Config

UPLOAD_FOLDER = Config.UPLOAD_FOLDER


def save_image(file, subfolder=""):
    folder_path = os.path.join(UPLOAD_FOLDER, subfolder)

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)

    original_filename = secure_filename(file.filename)
    name, ext = os.path.splitext(original_filename)
    unique_suffix = uuid.uuid4().hex
    filename = f"{name}_{unique_suffix}{ext}"

    filepath = os.path.join(folder_path, filename)
    file.save(filepath)

    return filepath.replace("\\", "/")


def get_model_by_name(model, field_name, value, error_message):
    instance = model.query.filter(getattr(model, field_name) == value).first()
    if not instance:
        raise ValueError(error_message)
    return instance


def remove_file_if_exists(file_path):
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
        except Exception as e:
            print(f"Ошибка при удалении файла {file_path}: {e}")


def is_valid_email(email):
    return bool(re.match(r"[^@]+@[^@]+\.[^@]+", email))


def send_async_email(app, msg):
    with app.app_context():
        try:
            mail.send(msg)
        except Exception as e:
            print(f"Ошибка при отправке email: {e}")


def send_email(subject, recipient, body, body_html=None, attachments=None):
    if not recipient or not is_valid_email(recipient):
        print(f"Попытка отправить email на невалидный адрес: {recipient}")
        return

    msg = Message(
        subject=subject,
        recipients=[recipient],
        body=body,
        html=body_html
    )

    if attachments:
        for attachment in attachments:
            if isinstance(attachment, tuple) and len(attachment) == 3:
                filename, content_bytes, mimetype = attachment
                data = BytesIO(content_bytes)
                msg.attach(filename, mimetype, data.read())
            elif isinstance(attachment, tuple) and len(attachment) == 2:
                filename, content_bytes = attachment
                data = BytesIO(content_bytes)
                msg.attach(filename, "text/csv; charset=utf-8", data.read())
            else:
                print(f"Некорректный формат вложения: {attachment}")

    threading.Thread(target=send_async_email, args=(current_app._get_current_object(), msg)).start()


def generate_reset_token(email, expires_sec=3600):
    s = URLSafeTimedSerializer(current_app.config["SECRET_KEY"])
    return s.dumps(email, salt='password-reset-salt')


def verify_reset_token(token, max_age=3600):
    s = URLSafeTimedSerializer(current_app.config["SECRET_KEY"])
    try:
        email = s.loads(token, salt='password-reset-salt', max_age=max_age)
    except Exception:
        return None
    return email


def to_str(value):
    if isinstance(value, bytes):
        return value.decode('utf-8')
    return str(value)


def format_datetime(value):
    if isinstance(value, datetime):
        return value.strftime('%d.%m.%Y %H:%M')
    return str(value)  # если не datetime, вернуть как есть


def generate_reservations_csv(reservations):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отменённые бронирования"

    headers = [
        ('reservation_id', 'ID бронирования'),
        ('full_name', 'ФИО'),
        ('email', 'Электронная почта'),
        ('phone_number', 'Телефон'),
        ('participants_count', 'Количество участников'),
        ('booked_at', 'Дата бронирования'),
        ('session_datetime', 'Время сессии'),
        ('excursion_title', 'Название экскурсии'),
        ('place', 'Место экскурсии'),
        ('total_cost', 'Общая стоимость'),
        ('is_paid', 'Оплачена'),
        ('is_cancelled', 'Отменена'),
    ]
    ws.append([h[1] for h in headers])

    for r in reservations:
        ws.append([
            str(r.get('reservation_id', '')),
            str(r.get('full_name', '')),
            str(r.get('email', '')),
            str(r.get('phone_number', '')),
            str(r.get('participants_count', '')),
            format_datetime(r.get('booked_at', '')),
            format_datetime(r.get('session_datetime', '')),
            str(r.get('excursion_title', '')),
            str(r.get('place', '')),
            str(r.get('total_cost', '')),
            'Да' if r.get('is_paid') else 'Нет',
            'Да' if r.get('is_cancelled') else 'Нет',
        ])

    for col_idx, column_cells in enumerate(ws.columns, 1):
        max_length = max((len(str(cell.value)) if cell.value else 0) for cell in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
